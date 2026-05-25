#!/usr/bin/env python3
"""extract_xlsx.py — Spreadsheet text extractor for translation pipelines.

Iterates all sheets and non-empty cells. Each cell gets a [Pnnn] tag
with location: {sheet, row, col}. Skips formula cells (no extractable
text content).
"""

import argparse
import json
import os


def parse_args():
    parser = argparse.ArgumentParser(
        description="Extract translatable text from an XLSX file."
    )
    parser.add_argument("--source", required=True)
    parser.add_argument("--extraction-output", required=True)
    parser.add_argument("--tagged-output", required=True)
    parser.add_argument("--report-output", required=True)
    return parser.parse_args()


def extract_from_xlsx(source_path):
    """Extract cell text from all sheets.

    Returns (paragraphs, warnings).
    Each cell: type="table", location={sheet, row, col, sheet_name}.
    """
    import openpyxl

    paragraphs = []
    warnings = []

    wb = openpyxl.load_workbook(source_path, data_only=True)

    for sheet_idx, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]
        for row_cells in ws.iter_rows():
            for cell in row_cells:
                if cell.value is None:
                    continue
                text = str(cell.value).strip()
                if not text:
                    continue
                paragraphs.append({
                    "tag": f"P{len(paragraphs)}",
                    "type": "table",
                    "text": text,
                    "location": {
                        "sheet": sheet_idx,
                        "row": cell.row - 1,
                        "col": cell.column - 1,
                        "sheet_name": sheet_name,
                    },
                })

    wb.close()
    return paragraphs, warnings


def ensure_parent_dir(filepath):
    parent = os.path.dirname(filepath)
    if parent:
        os.makedirs(parent, exist_ok=True)


def main():
    args = parse_args()
    paragraphs, warnings = extract_from_xlsx(args.source)

    ensure_parent_dir(args.extraction_output)
    extraction = {"source": os.path.abspath(args.source), "paragraphs": paragraphs}
    with open(args.extraction_output, "w", encoding="utf-8") as f:
        json.dump(extraction, f, ensure_ascii=False, indent=2)

    ensure_parent_dir(args.tagged_output)
    with open(args.tagged_output, "w", encoding="utf-8") as f:
        for p in paragraphs:
            f.write(f"[{p['tag']}] {p['text']}\n")

    ensure_parent_dir(args.report_output)
    report = {
        "paragraph_count": len(paragraphs),
        "table_count": len(paragraphs),
        "warnings": warnings,
    }
    with open(args.report_output, "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)

    print(f"Extraction complete: {report['paragraph_count']} cells.")


if __name__ == "__main__":
    main()
